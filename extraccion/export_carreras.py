#!/usr/bin/env python3
"""
export_carreras.py — Referencia determinista para clasificar y ordenar las
carreras de la USM a partir de planes_carreras.json.

Es la ÚNICA fuente de verdad para:

  1. Clasificar cada carrera en una categoría:
       Técnica | Profesional | Licenciatura | Magíster | Doctorado
     y descartar todo lo que no es carrera de grado (cursos, diplomados,
     programas, nivelación, talleres, postítulos, etc.).

  2. Determinar "Malla Nueva" vs "Malla Antigua" por plan:
       se ordenan las CLAVES de plan (plan_id, numérico y monotónico) de forma
       ascendente; la mayor es "Malla Nueva", el resto "Malla Antigua".
       Un plan único = "Malla Nueva". Sin planes = "Sin dato".

  3. Ordenar de forma canónica:
       sede -> jornada -> categoría -> carrera -> código -> mención -> plan.

Entrada : src/lib/data/planes_carreras.json  (generado por resources/generar.py)
Salida  : CSV (siempre, stdlib) y XLSX (opcional, requiere openpyxl).

Uso:
    python resources/export_carreras.py --csv carreras_usm.csv
    python resources/export_carreras.py --csv carreras_usm.csv --xlsx carreras_usm.xlsx
"""

import argparse
import csv
import json
import re
import sys
from collections import OrderedDict

# ---------------------------------------------------------------------------
# Constantes de dominio (orden canónico y reglas de clasificación).
# ---------------------------------------------------------------------------
SEDE_ORDER = [
    "Casa Central Valparaíso",
    "Santiago San Joaquín",
    "Santiago Vitacura",
    "Santiago (Vesp. JMC)",
    "Viña del Mar",
    "Concepción",
]

JORNADA_ORDER = ["Diurna", "Vespertina", "A Distancia"]

CATEGORIA_ORDER = ["Técnica", "Profesional", "Licenciatura", "Magíster", "Doctorado"]

# Nombres que SIEMPRE se descartan, aunque coincidan con un prefijo válido.
EXCLUIDAS_EXPLICITAS = {
    "Formación de Piloto Comercial",  # programa de formación, no título propio
}

# Nombres profesionales exactos que no empiezan con un prefijo reconocible.
PROFESIONALES_EXACTAS = {
    "Arquitectura",
    "Construcción Civil",
    "Químico",
    "Piloto Comercial",
}

# Prefijos de clasificación. El orden de evaluación es el orden de los `if`
# dentro de `clasificar` (importante por la precedencia).
RE_DOCTORADO = re.compile(r"^(Doc\.|Doctorado)", re.I)
RE_MAGISTER = re.compile(r"^(Mag\.|Magíster|Magister|Master)", re.I)
RE_LICENCIATURA = re.compile(r"^(Lic\.|Licenciatura)", re.I)
RE_TECNICA = re.compile(r"^(Téc\.|Tec\.|Técnico)", re.I)
RE_PROFESIONAL = re.compile(r"^(Ing\.|Ingeniería|Ingenieria|I\.Civil|I\.Ejec)", re.I)

# Valores de mención que no se reportan en la columna "Mención/Especialidad".
SIN_MENCION = {"sin mención", "sin mencion"}

HEADERS = [
    "Sede", "Jornada", "Código", "Carrera", "Categoría",
    "Mención / Especialidad", "Nº de Plan", "Tipo de Malla", "Nº Semestres",
]


def clasificar(nombre):
    """Devuelve la categoría de una carrera o None si debe descartarse."""
    n = (nombre or "").strip()
    if n in EXCLUIDAS_EXPLICITAS:
        return None
    if RE_DOCTORADO.match(n):
        return "Doctorado"
    if RE_MAGISTER.match(n):
        return "Magíster"
    if RE_LICENCIATURA.match(n):
        return "Licenciatura"
    if RE_TECNICA.match(n):
        return "Técnica"
    if RE_PROFESIONAL.match(n):
        return "Profesional"
    if n in PROFESIONALES_EXACTAS:
        return "Profesional"
    return None


def _es_sin_mencion(nombre):
    return (nombre or "").strip().lower() in SIN_MENCION


def _indice(secuencia, valor):
    try:
        return secuencia.index(valor)
    except ValueError:
        return len(secuencia)


def build_rows(carreras):
    """Convierte el contenido crudo de planes_carreras.json en filas ordenadas.

    Devuelve (rows, descartadas) donde:
      - rows       : lista de dicts con las columnas de HEADERS.
      - descartadas: lista de nombres de carrera que no son carrera de grado.
    """
    filas = []
    descartadas = []

    for c in carreras:
        sede = c.get("sede", "")
        jornada = c.get("jornada", "")
        codigo = c.get("código", "")
        nombre = c.get("nombre", "")
        categoria = clasificar(nombre)

        if categoria is None:
            descartadas.append((nombre or "").strip())
            continue

        menciones = c.get("menciones/especialidades", {}) or {}

        # Agrupar planes por plan_id (clave). Menciones distintas pueden
        # compartir el mismo plan_id; se deduplican aquí.
        planes = OrderedDict()
        for m in menciones.values():
            mnombre = m.get("nombre", "Sin mención")
            for plan_id, p in (m.get("planes", {}) or {}).items():
                if plan_id not in planes:
                    planes[plan_id] = {
                        "plan": p.get("plan", ""),
                        "semestres": len(p.get("malla", []) or []),
                        "menciones": set(),
                    }
                planes[plan_id]["menciones"].add(mnombre)

        if not planes:
            filas.append({
                "sede": sede, "jornada": jornada, "código": codigo,
                "carrera": nombre, "categoría": categoria, "mención": "",
                "plan": "", "tipo": "Sin dato", "semestres": "",
            })
            continue

        ids_ordenados = sorted(planes.keys(), key=lambda k: int(k))
        id_nuevo = ids_ordenados[-1]

        for plan_id in ids_ordenados:
            info = planes[plan_id]
            tipo = "Malla Nueva" if plan_id == id_nuevo else "Malla Antigua"
            menciones_reales = sorted(
                m for m in info["menciones"] if not _es_sin_mencion(m)
            )
            filas.append({
                "sede": sede, "jornada": jornada, "código": codigo,
                "carrera": nombre, "categoría": categoria,
                "mención": ", ".join(menciones_reales),
                "plan": str(info["plan"]),
                "tipo": tipo,
                "semestres": info["semestres"],
            })

    filas.sort(key=lambda r: (
        _indice(SEDE_ORDER, r["sede"]),
        _indice(JORNADA_ORDER, r["jornada"]),
        _indice(CATEGORIA_ORDER, r["categoría"]),
        r["carrera"].lower(),
        r["código"],
        r["mención"].lower(),
        0 if r["tipo"] == "Malla Nueva" else 1,
        r["plan"],
    ))

    return filas, descartadas


def _filas_a_valores(filas):
    return [
        [r["sede"], r["jornada"], r["código"], r["carrera"], r["categoría"],
         r["mención"], r["plan"], r["tipo"], r["semestres"]]
        for r in filas
    ]


def to_csv(filas, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(_filas_a_valores(filas))
    return path


def to_xlsx(filas, path):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[aviso] openpyxl no instalado; se omite XLSX.", file=sys.stderr)
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Carreras USM por Sede"
    ws.append(HEADERS)
    for vals in _filas_a_valores(filas):
        ws.append(vals)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(filas) + 1}"

    for col, w in enumerate([24, 12, 10, 40, 13, 28, 22, 16, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(path)
    return path


def main(argv=None):
    parser = argparse.ArgumentParser(description="Exporta carreras USM clasificadas y ordenadas.")
    parser.add_argument("--input", default="src/lib/data/planes_carreras.json",
                        help="Ruta a planes_carreras.json (default: %(default)s)")
    parser.add_argument("--csv", help="Ruta del archivo CSV de salida (recomendado para Git)")
    parser.add_argument("--xlsx", help="Ruta del archivo XLSX de salida (opcional)")
    args = parser.parse_args(argv)

    with open(args.input, encoding="utf-8") as f:
        carreras = json.load(f)

    filas, descartadas = build_rows(carreras)

    if args.csv:
        print("CSV ->", to_csv(filas, args.csv))
    if args.xlsx:
        out = to_xlsx(filas, args.xlsx)
        if out:
            print("XLSX ->", out)

    print(f"Filas de grado: {len(filas)}")
    print(f"Descartadas: {len(descartadas)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
